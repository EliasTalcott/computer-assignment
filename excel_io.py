#!/usr/bin/env python3

import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side

###
## Read class list from Excel file
###
def read_class_lists(infile):
    wb = openpyxl.load_workbook(filename = infile)
    sheet = wb.active
    lis = []
    for i in range(sheet.max_row - 1):
        lis.append(sheet.cell(row = i + 2, column = 1).value)
    return lis

###
## Write computer assignments to Excel file
###
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

def write_computer_assignments(outfile, lists, classes):
    # Open new workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "ComputerAssignments"

    # Formatting information
    numColsUsed = 2
    numRowsUsed = 4
    numClasses = len(classes)
    maxClassSize = 28

    # Format sheet
    sheet.merge_cells(start_row = 1, start_column = 1, end_row = 1, end_column = numClasses + 1)
    sheet.cell(column = 1, row = 1, value = "Computer Assignments").font = Font(size = "18", name = "Arial", bold = True)
    for row in range(maxClassSize + 3):
        sheet.row_dimensions[row + 1].height = 26.3
        if (row + 1) >= 4:
            sheet.cell(column = 1, row = row + 1, value = row - 2)
    for col in range(numClasses + 1):
        if (col + 1) == 1:
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col + 1)].width = 3.11
        else:
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col + 1)].width = 13.83
    for col in sheet:
        for cell in col:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrapText=True)
            cell.border = thin_border

    # Write assignment data
    for i, lis in enumerate(lists):
        sheet.cell(column = i + numColsUsed, row = 3, value = classes[i]).font = Font(size = "10", name = "Arial", bold = True)
        for j, student in enumerate(lis):
            sheet.cell(column = i + numColsUsed, row = j + numRowsUsed, value = " ".join(reversed(student.split(", ")))).font = Font(size = "10", name = "Arial")

    # Save workbook
    wb.save(filename = outfile)